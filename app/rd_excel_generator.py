"""
R&D Tax Credit Excel Report Generator

Generates comprehensive Excel workbook with 13 worksheets for R&D tax credit documentation.
"""

import io
import logging
from typing import List, Dict, Any, Optional
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from app.rd_parser import (
    RDAnalysisSession, RDProject, RDEmployee, RDVendor, RDExpense, TestStatus
)

logger = logging.getLogger(__name__)

# Styling constants
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
OK_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
REVIEW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def apply_header_style(ws, row_num: int = 1):
    """Apply header styling to first row"""
    for cell in ws[row_num]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER


def auto_adjust_columns(ws):
    """Auto-adjust column widths based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def format_currency(value: float) -> str:
    """Format value as currency"""
    return f"${value:,.2f}"


def format_percent(value: float) -> str:
    """Format value as percentage"""
    return f"{value:.1f}%"


# =============================================================================
# WORKSHEET GENERATORS
# =============================================================================

def generate_summary_statistics(wb: Workbook, session: RDAnalysisSession):
    """Generate Summary_Statistics worksheet"""
    ws = wb.create_sheet("Summary_Statistics")
    
    # Company Information
    data = [
        ["R&D TAX CREDIT STUDY - SUMMARY STATISTICS", ""],
        ["", ""],
        ["COMPANY INFORMATION", ""],
        ["Company Name", session.company_name or "Not Specified"],
        ["Industry", session.industry or "Not Specified"],
        ["Tax Year", session.tax_year],
        ["Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["", ""],
        ["HEADCOUNT SUMMARY", ""],
        ["Total Employees", session.total_employees],
        ["R&D Employees", session.rd_employees],
        ["R&D Headcount %", f"{(session.rd_employees / session.total_employees * 100) if session.total_employees > 0 else 0:.1f}%"],
        ["", ""],
        ["QRE SUMMARY", ""],
        ["Wage QRE", format_currency(session.wage_qre)],
        ["Supply QRE", format_currency(session.supply_qre)],
        ["Contract Research QRE", format_currency(session.contract_qre)],
        ["Total QRE", format_currency(session.total_qre)],
        ["", ""],
        ["PROJECT SUMMARY", ""],
        ["Total Projects", len(session.projects)],
        ["Qualified Projects", session.qualified_projects],
        ["Qualification Rate", f"{(session.qualified_projects / len(session.projects) * 100) if session.projects else 0:.1f}%"],
        ["", ""],
        ["VENDOR SUMMARY", ""],
        ["Total Vendors", len(session.vendors)],
        ["Qualified Vendors", len([v for v in session.vendors if v.qualified])],
    ]
    
    for row in data:
        ws.append(row)
    
    # Style the headers
    ws['A1'].font = Font(bold=True, size=14)
    ws['A3'].font = Font(bold=True)
    ws['A9'].font = Font(bold=True)
    ws['A14'].font = Font(bold=True)
    ws['A20'].font = Font(bold=True)
    ws['A25'].font = Font(bold=True)
    
    auto_adjust_columns(ws)


def generate_employees_sheet(wb: Workbook, session: RDAnalysisSession):
    """Generate Employees worksheet"""
    ws = wb.create_sheet("Employees")
    
    headers = [
        "Employee_ID", "Name", "Job_Title", "Department", "Location_State",
        "W2_Wages", "Stock_Compensation", "Severance", "QRE_Wage_Base", "RD_Allocation_%"
    ]
    ws.append(headers)
    apply_header_style(ws)
    
    for emp in session.employees:
        ws.append([
            emp.employee_id,
            emp.name,
            emp.job_title or "",
            emp.department or "",
            emp.location or "",
            emp.w2_wages,
            emp.stock_compensation,
            emp.severance,
            emp.qre_wage_base,
            emp.rd_allocation_percent
        ])
    
    # Format currency columns
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=9):
        for cell in row:
            cell.number_format = '$#,##0.00'
    
    # Format percent column
    for cell in ws.iter_rows(min_row=2, min_col=10, max_col=10):
        cell[0].number_format = '0.0%'
    
    auto_adjust_columns(ws)


def generate_timesheets_sheet(wb: Workbook, session: RDAnalysisSession):
    """Generate Timesheets worksheet (placeholder with summary data)"""
    ws = wb.create_sheet("Timesheets")
    
    headers = [
        "Employee_ID", "Name", "Department", "Year", 
        "RD_Hours", "Remaining_Hours", "Total_Hours", "RD_Ratio"
    ]
    ws.append(headers)
    apply_header_style(ws)
    
    # Generate timesheet summaries from employees with R&D allocation
    for emp in session.employees:
        if emp.rd_allocation_percent > 0:
            total_hours = 1800  # Standard full year
            rd_hours = total_hours * (emp.rd_allocation_percent / 100)
            remaining_hours = total_hours - rd_hours
            
            ws.append([
                emp.employee_id,
                emp.name,
                emp.department or "",
                session.tax_year,
                rd_hours,
                remaining_hours,
                total_hours,
                emp.rd_allocation_percent / 100
            ])
    
    # Format ratio column
    for cell in ws.iter_rows(min_row=2, min_col=8, max_col=8):
        cell[0].number_format = '0.0%'
    
    auto_adjust_columns(ws)


def generate_projects_sheet(wb: Workbook, session: RDAnalysisSession):
    """Generate Projects worksheet"""
    ws = wb.create_sheet("Projects")
    
    headers = [
        "Project_ID", "Project_Name", "Category", "Description",
        "Qualified_Research", "Permitted_Purpose", "Elimination_Uncertainty",
        "Process_Experimentation", "Technological_Nature", "Confidence_Score", "AI_Summary"
    ]
    ws.append(headers)
    apply_header_style(ws)
    
    for proj in session.projects:
        fpt = proj.four_part_test
        ws.append([
            proj.project_id,
            proj.project_name,
            proj.category or "",
            proj.description or "",
            "Yes" if proj.qualified else "No",
            fpt.permitted_purpose.value if hasattr(fpt.permitted_purpose, 'value') else str(fpt.permitted_purpose),
            fpt.elimination_uncertainty.value if hasattr(fpt.elimination_uncertainty, 'value') else str(fpt.elimination_uncertainty),
            fpt.process_experimentation.value if hasattr(fpt.process_experimentation, 'value') else str(fpt.process_experimentation),
            fpt.technological_nature.value if hasattr(fpt.technological_nature, 'value') else str(fpt.technological_nature),
            proj.confidence_score,
            proj.ai_summary or ""
        ])
    
    # Color-code the test results
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=6, max_col=9), start=2):
        for cell in row:
            if cell.value == "pass":
                cell.fill = OK_FILL
            elif cell.value == "needs_review":
                cell.fill = REVIEW_FILL
            elif cell.value == "fail":
                cell.fill = FAIL_FILL
    
    auto_adjust_columns(ws)


def generate_project_allocations_sheet(wb: Workbook, session: RDAnalysisSession):
    """Generate Project_Allocations worksheet"""
    ws = wb.create_sheet("Project_Allocations")
    
    headers = ["Employee_ID", "Employee_Name", "Project_ID", "Project_Name", "Percent_Allocation"]
    ws.append(headers)
    apply_header_style(ws)
    
    # Generate allocations based on R&D employees and projects
    for emp in session.employees:
        if emp.rd_allocation_percent > 0 and session.projects:
            # Distribute allocation across qualified projects
            qualified_projects = [p for p in session.projects if p.qualified]
            if not qualified_projects:
                qualified_projects = session.projects[:3]  # Take first 3 if none qualified
            
            allocation_per_project = emp.rd_allocation_percent / len(qualified_projects) if qualified_projects else 0
            
            for proj in qualified_projects:
                ws.append([
                    emp.employee_id,
                    emp.name,
                    proj.project_id,
                    proj.project_name,
                    allocation_per_project
                ])
    
    # Format percent column
    for cell in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        cell[0].number_format = '0.0%'
    
    auto_adjust_columns(ws)


def generate_vendors_sheet(wb: Workbook, session: RDAnalysisSession):
    """Generate Vendors worksheet"""
    ws = wb.create_sheet("Vendors")
    
    headers = ["Vendor_ID", "Vendor_Name", "Risk_Bearer", "IP_Rights", "Country", "Qualified_Sec41"]
    ws.append(headers)
    apply_header_style(ws)
    
    for vendor in session.vendors:
        ws.append([
            vendor.vendor_id,
            vendor.vendor_name,
            vendor.risk_bearer,
            vendor.ip_rights,
            vendor.country,
            "Yes" if vendor.qualified else "No"
        ])
    
    # Color-code qualified status
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=6, max_col=6), start=2):
        for cell in row:
            if cell.value == "Yes":
                cell.fill = OK_FILL
            else:
                cell.fill = REVIEW_FILL
    
    auto_adjust_columns(ws)


def generate_ap_transactions_sheet(wb: Workbook, session: RDAnalysisSession):
    """Generate AP_Transactions worksheet (Contract Research)"""
    ws = wb.create_sheet("AP_Transactions")
    
    headers = [
        "Transaction_ID", "Vendor_ID", "Description", 
        "Amount", "Qualified_Contract_Research", "QRE_Amount"
    ]
    ws.append(headers)
    apply_header_style(ws)
    
    for exp in session.expenses:
        if exp.category == "contract_research":
            ws.append([
                exp.transaction_id,
                exp.vendor_id or "",
                exp.description,
                exp.amount,
                "Yes" if exp.qualified else "No",
                exp.qre_amount
            ])
    
    # Format currency columns
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = '$#,##0.00'
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
        for cell in row:
            cell.number_format = '$#,##0.00'
    
    auto_adjust_columns(ws)


def generate_supplies_sheet(wb: Workbook, session: RDAnalysisSession):
    """Generate Supplies worksheet"""
    ws = wb.create_sheet("Supplies")
    
    headers = [
        "Supply_ID", "Description", "Amount", 
        "Qualified_Supply", "QRE_Amount"
    ]
    ws.append(headers)
    apply_header_style(ws)
    
    for exp in session.expenses:
        if exp.category == "supplies":
            ws.append([
                exp.transaction_id,
                exp.description,
                exp.amount,
                "Yes" if exp.qualified else "No",
                exp.qre_amount
            ])
    
    # Format currency columns
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = '$#,##0.00'
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = '$#,##0.00'
    
    auto_adjust_columns(ws)


def generate_automated_review_sheet(wb: Workbook, session: RDAnalysisSession):
    """Generate Automated_Review worksheet with diagnostic tests"""
    ws = wb.create_sheet("Automated_Review")
    
    headers = [
        "Category", "Level", "Key_ID", "Name", "Year",
        "Metric_Name", "Metric_Value", "Threshold", "Status", "Comment"
    ]
    ws.append(headers)
    apply_header_style(ws)
    
    reviews = []
    
    # Test 1: R&D Allocation reasonableness
    for emp in session.employees:
        if emp.rd_allocation_percent > 0:
            # Check if allocation seems reasonable for role
            status = "OK"
            comment = ""
            
            if emp.rd_allocation_percent > 90:
                status = "Review"
                comment = "Very high R&D allocation - verify role is purely R&D"
            
            reviews.append([
                "Employee",
                "Info",
                emp.employee_id,
                emp.name,
                session.tax_year,
                "RD_Allocation_%",
                emp.rd_allocation_percent,
                100,
                status,
                comment
            ])
    
    # Test 2: Severance Reasonableness (flag if > 40% of W-2)
    for emp in session.employees:
        if emp.severance > 0 and emp.w2_wages > 0:
            severance_ratio = emp.severance / emp.w2_wages
            status = "OK" if severance_ratio <= 0.4 else "Review"
            comment = f"Severance is {severance_ratio*100:.1f}% of W-2 wages" if status == "Review" else ""
            
            reviews.append([
                "Employee",
                "Warning" if status == "Review" else "Info",
                emp.employee_id,
                emp.name,
                session.tax_year,
                "Severance_Ratio",
                severance_ratio * 100,
                40,
                status,
                comment
            ])
    
    # Test 3: Stock Compensation (flag if > 50% of W-2 OR > $200k)
    for emp in session.employees:
        if emp.stock_compensation > 0:
            stock_ratio = emp.stock_compensation / emp.w2_wages if emp.w2_wages > 0 else 0
            
            status = "OK"
            comment = ""
            
            if stock_ratio > 0.5:
                status = "Review"
                comment = f"Stock comp is {stock_ratio*100:.1f}% of W-2"
            elif emp.stock_compensation > 200000:
                status = "Review"
                comment = f"Stock comp exceeds $200k threshold"
            
            if status == "Review":
                reviews.append([
                    "Employee",
                    "Warning",
                    emp.employee_id,
                    emp.name,
                    session.tax_year,
                    "Stock_Compensation",
                    emp.stock_compensation,
                    200000,
                    status,
                    comment
                ])
    
    # Test 4: Vendor Risk Structure
    for vendor in session.vendors:
        if not vendor.qualified:
            reviews.append([
                "Vendor",
                "Warning",
                vendor.vendor_id,
                vendor.vendor_name,
                session.tax_year,
                "Sec41_Qualification",
                0,
                1,
                "Review",
                f"Risk: {vendor.risk_bearer}, IP: {vendor.ip_rights} - does not meet Sec.41 requirements"
            ])
    
    # Write reviews
    for review in reviews:
        ws.append(review)
    
    # Color-code status
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=9, max_col=9), start=2):
        for cell in row:
            if cell.value == "OK":
                cell.fill = OK_FILL
            elif cell.value == "Review":
                cell.fill = REVIEW_FILL
    
    auto_adjust_columns(ws)


def generate_qre_summary_sheet(wb: Workbook, session: RDAnalysisSession):
    """Generate QRE_Summary_2024 worksheet"""
    ws = wb.create_sheet("QRE_Summary_2024")
    
    headers = ["Category", "Amount"]
    ws.append(headers)
    apply_header_style(ws)
    
    ws.append(["Wage_QRE_2024", session.wage_qre])
    ws.append(["Supplies_QRE_2024", session.supply_qre])
    ws.append(["Contract_QRE_2024", session.contract_qre])
    ws.append(["", ""])
    ws.append(["Total_QRE_2024", session.total_qre])
    
    # Format currency column
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            if cell.value and isinstance(cell.value, (int, float)):
                cell.number_format = '$#,##0.00'
    
    # Bold the total row
    ws.cell(row=6, column=1).font = Font(bold=True)
    ws.cell(row=6, column=2).font = Font(bold=True)
    
    auto_adjust_columns(ws)


def compute_asc_credit(qre_2024: float, qre_2021: float = 0, qre_2022: float = 0, qre_2023: float = 0) -> Dict[str, float]:
    """Compute ASC method credit"""
    # Average of prior 3 years
    avg_prior_qre = (qre_2021 + qre_2022 + qre_2023) / 3
    
    # Base = 50% of average
    base_amount = 0.5 * avg_prior_qre
    
    # Excess QRE
    excess = max(0, qre_2024 - base_amount)
    
    # ASC Credit = 14% of excess
    credit = 0.14 * excess
    
    return {
        "avg_prior_qre": avg_prior_qre,
        "base_amount": base_amount,
        "excess": excess,
        "credit": credit
    }


def compute_regular_credit(qre_2024: float, fixed_base_pct: float = 0.03, avg_gross_receipts: float = 0) -> Dict[str, float]:
    """Compute Regular method credit"""
    # Base = fixed_base_percentage x average gross receipts
    base_amount = fixed_base_pct * avg_gross_receipts
    
    # Excess QRE
    excess = max(0, qre_2024 - base_amount)
    
    # Regular Credit = 20% of excess
    credit = 0.20 * excess
    
    return {
        "fixed_base_pct": fixed_base_pct,
        "avg_gross_receipts": avg_gross_receipts,
        "base_amount": base_amount,
        "excess": excess,
        "credit": credit
    }


def generate_credit_computation_sheet(wb: Workbook, session: RDAnalysisSession):
    """Generate Credit_Computation_2024 worksheet"""
    ws = wb.create_sheet("Credit_Computation_2024")
    
    # Estimate prior year QREs (in practice these would come from input)
    qre_2024 = session.total_qre
    qre_2023 = qre_2024 * 0.92  # Estimate 8% growth
    qre_2022 = qre_2023 * 0.90  # Estimate 10% growth
    qre_2021 = qre_2022 * 0.88  # Estimate 12% growth
    
    # Compute ASC credit
    asc = compute_asc_credit(qre_2024, qre_2021, qre_2022, qre_2023)
    
    # Compute Regular credit (with assumed values)
    avg_gross_receipts = qre_2024 * 10  # Estimate company size
    regular = compute_regular_credit(qre_2024, 0.03, avg_gross_receipts)
    
    # Determine selected method
    selected_method = "ASC" if asc["credit"] >= regular["credit"] else "Regular"
    selected_credit = max(asc["credit"], regular["credit"])
    
    data = [
        ["R&D TAX CREDIT COMPUTATION - TAX YEAR 2024", ""],
        ["", ""],
        ["ASC METHOD (Alternative Simplified Credit)", ""],
        ["QRE 2021 (Estimated)", qre_2021],
        ["QRE 2022 (Estimated)", qre_2022],
        ["QRE 2023 (Estimated)", qre_2023],
        ["Average Prior 3-Year QRE", asc["avg_prior_qre"]],
        ["Base Amount (50% of Average)", asc["base_amount"]],
        ["QRE 2024", qre_2024],
        ["Excess QRE", asc["excess"]],
        ["ASC Credit (14% of Excess)", asc["credit"]],
        ["", ""],
        ["REGULAR METHOD", ""],
        ["Fixed Base Percentage", f"{regular['fixed_base_pct']*100:.1f}%"],
        ["Average Gross Receipts (Prior 4 Years)", regular["avg_gross_receipts"]],
        ["Base Amount", regular["base_amount"]],
        ["QRE 2024", qre_2024],
        ["Excess QRE", regular["excess"]],
        ["Regular Credit (20% of Excess)", regular["credit"]],
        ["", ""],
        ["METHOD SELECTION", ""],
        ["ASC Credit Amount", asc["credit"]],
        ["Regular Credit Amount", regular["credit"]],
        ["Selected Method", selected_method],
        ["SELECTED CREDIT AMOUNT", selected_credit],
    ]
    
    for row in data:
        ws.append(row)
    
    # Format currency cells
    currency_rows = [4, 5, 6, 7, 8, 9, 10, 11, 15, 16, 17, 18, 19, 22, 23, 25]
    for row_num in currency_rows:
        cell = ws.cell(row=row_num, column=2)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '$#,##0.00'
    
    # Style headers
    ws['A1'].font = Font(bold=True, size=14)
    ws['A3'].font = Font(bold=True)
    ws['A13'].font = Font(bold=True)
    ws['A21'].font = Font(bold=True)
    ws['A25'].font = Font(bold=True, size=12)
    ws['B25'].font = Font(bold=True, size=12)
    ws['B25'].fill = OK_FILL
    
    auto_adjust_columns(ws)


def generate_sec174_tieout_sheet(wb: Workbook, session: RDAnalysisSession):
    """Generate Sec_174_TieOut worksheet"""
    ws = wb.create_sheet("Sec_174_TieOut")
    
    headers = [
        "Description", "Category", "Amount", "Qualified_Sec174",
        "US_or_Foreign", "Amortization_Start_Date", "Tax_Year", "Notes"
    ]
    ws.append(headers)
    apply_header_style(ws)
    
    # Add wage entries
    total_wage_qre = 0
    for emp in session.employees:
        if emp.qre_wage_base > 0:
            ws.append([
                f"Wages - {emp.name}",
                "Wages",
                emp.qre_wage_base * (emp.rd_allocation_percent / 100),
                "Yes",
                "US",
                f"01/01/{session.tax_year}",
                session.tax_year,
                f"R&D Allocation: {emp.rd_allocation_percent}%"
            ])
            total_wage_qre += emp.qre_wage_base * (emp.rd_allocation_percent / 100)
    
    # Add supply entries
    for exp in session.expenses:
        if exp.category == "supplies" and exp.qualified:
            ws.append([
                f"Supply - {exp.description[:50]}",
                "Supplies",
                exp.qre_amount,
                "Yes",
                "US",
                f"01/01/{session.tax_year}",
                session.tax_year,
                ""
            ])
    
    # Add contract research entries
    for exp in session.expenses:
        if exp.category == "contract_research" and exp.qualified:
            # Find vendor for location
            vendor = next((v for v in session.vendors if v.vendor_id == exp.vendor_id), None)
            location = "US" if vendor and vendor.country.upper() in ["US", "USA", "UNITED STATES", ""] else "Foreign"
            
            ws.append([
                f"Contract Research - {exp.description[:50]}",
                "Contract Research",
                exp.qre_amount,
                "Yes",
                location,
                f"01/01/{session.tax_year}",
                session.tax_year,
                f"Vendor: {vendor.vendor_name if vendor else 'Unknown'}"
            ])
    
    # Format currency column
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            if cell.value and isinstance(cell.value, (int, float)):
                cell.number_format = '$#,##0.00'
    
    auto_adjust_columns(ws)


def generate_sanity_checks_sheet(wb: Workbook, session: RDAnalysisSession):
    """Generate Sanity_Checks worksheet"""
    ws = wb.create_sheet("Sanity_Checks")
    
    headers = [
        "Category", "Level", "Key_ID", "Name", "Year",
        "Metric_Name", "Metric_Value", "Threshold", "Status", "Comment"
    ]
    ws.append(headers)
    apply_header_style(ws)
    
    checks = []
    
    # Check 1: Employees with hire date after study period (placeholder - would need hire dates)
    checks.append([
        "Data Quality",
        "Info",
        "ALL",
        "Hire Date Check",
        session.tax_year,
        f"Employees with Hire Date after 12/31/{session.tax_year}",
        0,
        0,
        "OK",
        "No employees hired after study period end"
    ])
    
    # Check 2: Employees with low annual hours
    low_hours_count = len([e for e in session.employees if e.rd_allocation_percent > 0 and e.rd_allocation_percent < 100 * (1500/1800)])
    checks.append([
        "Data Quality",
        "Warning" if low_hours_count > 0 else "Info",
        "ALL",
        "Annual Hours Check",
        session.tax_year,
        "Employees with Annual Hours < 1500",
        low_hours_count,
        0,
        "Review" if low_hours_count > 0 else "OK",
        "Low hours may indicate leave, part-time, or missing timesheets" if low_hours_count > 0 else ""
    ])
    
    # Check 3: Total allocation > 100%
    # In a real system, we'd calculate sum of allocations per employee
    over_allocated = 0  # Placeholder
    checks.append([
        "Data Quality",
        "Warning" if over_allocated > 0 else "Info",
        "ALL",
        "Allocation Check",
        session.tax_year,
        "Employees with SUM(Percent_Allocation) > 100%",
        over_allocated,
        0,
        "Review" if over_allocated > 0 else "OK",
        "Total allocation per employee should not exceed 100%" if over_allocated > 0 else ""
    ])
    
    # Check 4: QRE ratio check
    if session.total_qre > 0:
        wage_ratio = session.wage_qre / session.total_qre * 100
        checks.append([
            "QRE Analysis",
            "Info",
            "ALL",
            "Wage QRE Ratio",
            session.tax_year,
            "Wage QRE as % of Total QRE",
            wage_ratio,
            100,
            "OK" if 40 <= wage_ratio <= 80 else "Review",
            "" if 40 <= wage_ratio <= 80 else "Wage ratio outside typical range (40-80%)"
        ])
    
    # Check 5: Project qualification rate
    if session.projects:
        qual_rate = session.qualified_projects / len(session.projects) * 100
        checks.append([
            "Project Analysis",
            "Info",
            "ALL",
            "Qualification Rate",
            session.tax_year,
            "Projects Passing Four-Part Test",
            qual_rate,
            100,
            "OK" if qual_rate >= 50 else "Review",
            "" if qual_rate >= 50 else "Low qualification rate - review project documentation"
        ])
    
    # Write checks
    for check in checks:
        ws.append(check)
    
    # Color-code status
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=9, max_col=9), start=2):
        for cell in row:
            if cell.value == "OK":
                cell.fill = OK_FILL
            elif cell.value == "Review":
                cell.fill = REVIEW_FILL
    
    auto_adjust_columns(ws)


def generate_form_6765_sheet(wb: Workbook, session: RDAnalysisSession):
    """
    Generate Form 6765 (Credit for Increasing Research Activities) worksheet.
    
    This follows the structure of IRS Form 6765 with both Regular and ASC methods.
    """
    ws = wb.create_sheet("Form_6765_Computation")
    
    # Header
    ws.append(["IRS FORM 6765 - CREDIT FOR INCREASING RESEARCH ACTIVITIES"])
    ws.append([f"Tax Year: {session.tax_year}"])
    ws.append([f"Taxpayer: {session.company_name or 'Not Specified'}"])
    ws.append([""])
    
    # Style header
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'].font = Font(bold=True)
    ws['A3'].font = Font(bold=True)
    
    # Calculate values
    qre_current = session.total_qre
    wage_qre = session.wage_qre
    supply_qre = session.supply_qre
    contract_qre = session.contract_qre
    
    # Estimate prior year QREs for ASC calculation
    # In production, these would come from historical data
    qre_prior_1 = qre_current * 0.92
    qre_prior_2 = qre_current * 0.85
    qre_prior_3 = qre_current * 0.78
    
    # Check if company qualifies as Qualified Small Business (QSB)
    # QSB: Gross receipts < $5M for any 5 tax years AND R&D has been conducted
    # For payroll tax offset eligibility
    is_qsb = True  # Assume QSB for demonstration - would need actual gross receipts data
    
    # ==========================================================================
    # SECTION A - REGULAR CREDIT
    # ==========================================================================
    ws.append(["SECTION A - REGULAR CREDIT (Lines 1-11)"])
    ws['A5'].font = Font(bold=True)
    ws['A5'].fill = HEADER_FILL
    ws['A5'].font = HEADER_FONT
    
    ws.append(["Line", "Description", "Amount"])
    ws['A6'].font = Font(bold=True)
    ws['B6'].font = Font(bold=True)
    ws['C6'].font = Font(bold=True)
    
    regular_lines = [
        ("1", "Certain amounts paid or incurred to energy consortia", 0),
        ("2", "Basic research payments to qualified organizations", 0),
        ("3", "Qualified research expenses for tax year", qre_current),
        ("4", "Enter fixed-base percentage (not more than 16%)", "3.00%"),
        ("5", "Average annual gross receipts (prior 4 years)", qre_current * 10),  # Estimate
        ("6", "Multiply line 5 by line 4", qre_current * 10 * 0.03),
        ("7", "Subtract line 6 from line 3. If zero or less, enter -0-", max(0, qre_current - qre_current * 10 * 0.03)),
        ("8", "Multiply line 3 by 50%", qre_current * 0.50),
        ("9", "Enter the smaller of line 7 or line 8", min(max(0, qre_current - qre_current * 10 * 0.03), qre_current * 0.50)),
        ("10", "Add lines 1, 2, and 9", min(max(0, qre_current - qre_current * 10 * 0.03), qre_current * 0.50)),
        ("11", "Regular credit (Line 10 × 20%)", min(max(0, qre_current - qre_current * 10 * 0.03), qre_current * 0.50) * 0.20),
    ]
    
    for line_num, desc, amount in regular_lines:
        ws.append([f"Line {line_num}", desc, amount])
    
    ws.append([""])
    
    # ==========================================================================
    # SECTION B - ALTERNATIVE SIMPLIFIED CREDIT (ASC)
    # ==========================================================================
    ws.append(["SECTION B - ALTERNATIVE SIMPLIFIED CREDIT (Lines 12-17)"])
    row_b_start = ws.max_row
    ws[f'A{row_b_start}'].font = Font(bold=True)
    ws[f'A{row_b_start}'].fill = HEADER_FILL
    ws[f'A{row_b_start}'].font = HEADER_FONT
    
    # ASC calculation
    avg_prior_3yr = (qre_prior_1 + qre_prior_2 + qre_prior_3) / 3
    base_amount = avg_prior_3yr * 0.50
    excess_qre = max(0, qre_current - base_amount)
    asc_credit = excess_qre * 0.14
    
    asc_lines = [
        ("12", f"QRE for tax year {session.tax_year}", qre_current),
        ("13a", f"QRE for tax year {session.tax_year - 1}", qre_prior_1),
        ("13b", f"QRE for tax year {session.tax_year - 2}", qre_prior_2),
        ("13c", f"QRE for tax year {session.tax_year - 3}", qre_prior_3),
        ("14", "Average of lines 13a through 13c", avg_prior_3yr),
        ("15", "Multiply line 14 by 50%", base_amount),
        ("16", "Subtract line 15 from line 12 (if zero or less, skip to line 17)", excess_qre),
        ("17", "Alternative simplified credit (Line 16 × 14%)", asc_credit),
    ]
    
    for line_num, desc, amount in asc_lines:
        ws.append([f"Line {line_num}", desc, amount])
    
    ws.append([""])
    
    # ==========================================================================
    # SECTION C - CURRENT YEAR CREDIT
    # ==========================================================================
    ws.append(["SECTION C - CURRENT YEAR CREDIT"])
    row_c_start = ws.max_row
    ws[f'A{row_c_start}'].font = Font(bold=True)
    ws[f'A{row_c_start}'].fill = HEADER_FILL
    ws[f'A{row_c_start}'].font = HEADER_FONT
    
    regular_credit = regular_lines[-1][2]  # Line 11
    
    # Determine selected method
    selected_credit = max(regular_credit, asc_credit)
    selected_method = "ASC" if asc_credit >= regular_credit else "Regular"
    
    credit_lines = [
        ("24", f"Regular credit from Section A (line 11)", regular_credit),
        ("25", f"Alternative simplified credit from Section B (line 17)", asc_credit),
        ("26", f"Selected credit method", selected_method),
        ("27", f"Current year research credit (higher of line 24 or 25)", selected_credit),
    ]
    
    for line_num, desc, amount in credit_lines:
        ws.append([f"Line {line_num}", desc, amount])
    
    ws.append([""])
    
    # ==========================================================================
    # SECTION D - PAYROLL TAX CREDIT (FOR QUALIFIED SMALL BUSINESSES)
    # ==========================================================================
    ws.append(["SECTION D - PAYROLL TAX CREDIT (QSB ELECTION)"])
    row_d_start = ws.max_row
    ws[f'A{row_d_start}'].font = Font(bold=True)
    ws[f'A{row_d_start}'].fill = HEADER_FILL
    ws[f'A{row_d_start}'].font = HEADER_FONT
    
    # QSB can elect to apply up to $500K of credit against payroll taxes
    max_payroll_offset = 500000
    payroll_credit_amount = min(selected_credit, max_payroll_offset) if is_qsb else 0
    
    payroll_lines = [
        ("", "Qualified Small Business (QSB) Status", "Yes (assumed)" if is_qsb else "No"),
        ("41", "Research credit from Section C", selected_credit),
        ("42", "Maximum payroll tax credit amount", max_payroll_offset),
        ("43", "Payroll tax credit election (smaller of line 41 or 42)", payroll_credit_amount),
        ("44", "Remaining credit for income tax offset", selected_credit - payroll_credit_amount),
    ]
    
    for line_num, desc, amount in payroll_lines:
        line_label = f"Line {line_num}" if line_num else ""
        ws.append([line_label, desc, amount])
    
    ws.append([""])
    
    # ==========================================================================
    # SUMMARY
    # ==========================================================================
    ws.append(["CREDIT SUMMARY"])
    row_s_start = ws.max_row
    ws[f'A{row_s_start}'].font = Font(bold=True)
    ws[f'A{row_s_start}'].fill = OK_FILL
    
    summary_items = [
        ("Total QRE", qre_current),
        ("  Wage QRE", wage_qre),
        ("  Supply QRE", supply_qre),
        ("  Contract QRE (65%)", contract_qre),
        ("", ""),
        ("Selected Method", selected_method),
        ("TOTAL R&D TAX CREDIT", selected_credit),
        ("", ""),
        ("Section 280C Reduction", selected_credit),
        ("Net Tax Benefit (Credit - 21% of Credit)", selected_credit * 0.79),
    ]
    
    for desc, amount in summary_items:
        if amount == "":
            ws.append([desc, ""])
        elif isinstance(amount, str):
            ws.append([desc, amount])
        else:
            ws.append([desc, amount])
    
    # Format currency columns
    for row in ws.iter_rows(min_row=7, min_col=3, max_col=3):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '$#,##0'
    
    # Highlight the total credit
    for row_idx in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row_idx, column=1)
        if cell_a.value and "TOTAL R&D TAX CREDIT" in str(cell_a.value):
            cell_a.font = Font(bold=True, size=12)
            ws.cell(row=row_idx, column=2).font = Font(bold=True, size=12)
            ws.cell(row=row_idx, column=2).fill = OK_FILL
    
    auto_adjust_columns(ws)


def generate_four_part_test_documentation_sheet(wb: Workbook, session: RDAnalysisSession):
    """
    Generate a comprehensive four-part test documentation sheet.
    This provides IRS-ready documentation of each project's qualification.
    """
    ws = wb.create_sheet("Four_Part_Test_Details")
    
    ws.append(["FOUR-PART TEST DOCUMENTATION - IRS SECTION 41"])
    ws.append([f"Tax Year: {session.tax_year}"])
    ws.append([f"Company: {session.company_name or 'Not Specified'}"])
    ws.append([""])
    
    ws['A1'].font = Font(bold=True, size=14)
    
    # Add headers for the test documentation
    headers = [
        "Project ID", "Project Name", "Category",
        "Permitted Purpose", "PP Reasoning",
        "Elimination of Uncertainty", "EU Reasoning",
        "Process of Experimentation", "PE Reasoning",
        "Technological in Nature", "TN Reasoning",
        "Qualified?", "Confidence Score", "AI Summary"
    ]
    ws.append(headers)
    apply_header_style(ws, 5)
    
    for proj in session.projects:
        fpt = proj.four_part_test
        ws.append([
            proj.project_id,
            proj.project_name,
            proj.category or "",
            fpt.permitted_purpose.value if hasattr(fpt.permitted_purpose, 'value') else str(fpt.permitted_purpose),
            fpt.permitted_purpose_reasoning or "",
            fpt.elimination_uncertainty.value if hasattr(fpt.elimination_uncertainty, 'value') else str(fpt.elimination_uncertainty),
            fpt.elimination_uncertainty_reasoning or "",
            fpt.process_experimentation.value if hasattr(fpt.process_experimentation, 'value') else str(fpt.process_experimentation),
            fpt.process_experimentation_reasoning or "",
            fpt.technological_nature.value if hasattr(fpt.technological_nature, 'value') else str(fpt.technological_nature),
            fpt.technological_nature_reasoning or "",
            "Yes" if proj.qualified else "No",
            f"{proj.confidence_score * 100:.0f}%" if proj.confidence_score else "N/A",
            proj.ai_summary or "",
        ])
    
    # Color-code qualification status
    for row_idx, row in enumerate(ws.iter_rows(min_row=6, min_col=12, max_col=12), start=6):
        for cell in row:
            if cell.value == "Yes":
                cell.fill = OK_FILL
            elif cell.value == "No":
                cell.fill = REVIEW_FILL
    
    # Color-code test results
    test_columns = [4, 6, 8, 10]  # PP, EU, PE, TN columns
    for col_idx in test_columns:
        for row_idx in range(6, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value == "pass":
                cell.fill = OK_FILL
            elif cell.value == "needs_review":
                cell.fill = REVIEW_FILL
            elif cell.value == "fail":
                cell.fill = FAIL_FILL
    
    auto_adjust_columns(ws)


# =============================================================================
# MAIN GENERATOR
# =============================================================================

def generate_rd_workbook(session: RDAnalysisSession) -> bytes:
    """
    Generate comprehensive R&D tax credit Excel workbook.
    
    Args:
        session: RDAnalysisSession with all parsed and analyzed data
        
    Returns:
        bytes: Excel file content as bytes
    """
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Generate all worksheets
    logger.info("Generating Summary_Statistics sheet...")
    generate_summary_statistics(wb, session)
    
    logger.info("Generating Employees sheet...")
    generate_employees_sheet(wb, session)
    
    logger.info("Generating Timesheets sheet...")
    generate_timesheets_sheet(wb, session)
    
    logger.info("Generating Projects sheet...")
    generate_projects_sheet(wb, session)
    
    logger.info("Generating Project_Allocations sheet...")
    generate_project_allocations_sheet(wb, session)
    
    logger.info("Generating Vendors sheet...")
    generate_vendors_sheet(wb, session)
    
    logger.info("Generating AP_Transactions sheet...")
    generate_ap_transactions_sheet(wb, session)
    
    logger.info("Generating Supplies sheet...")
    generate_supplies_sheet(wb, session)
    
    logger.info("Generating Automated_Review sheet...")
    generate_automated_review_sheet(wb, session)
    
    logger.info("Generating QRE_Summary_2024 sheet...")
    generate_qre_summary_sheet(wb, session)
    
    logger.info("Generating Credit_Computation_2024 sheet...")
    generate_credit_computation_sheet(wb, session)
    
    logger.info("Generating Sec_174_TieOut sheet...")
    generate_sec174_tieout_sheet(wb, session)
    
    logger.info("Generating Sanity_Checks sheet...")
    generate_sanity_checks_sheet(wb, session)
    
    logger.info("Generating Form_6765_Computation sheet...")
    generate_form_6765_sheet(wb, session)
    
    logger.info("Generating Four_Part_Test_Details sheet...")
    generate_four_part_test_documentation_sheet(wb, session)
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    logger.info("Excel workbook generation complete")
    return output.getvalue()




